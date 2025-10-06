#!/usr/bin/env python3
"""
dLocal Pending Debits Processing Web Application
Upload transaction files and generate journal entries automatically
"""

from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import os
from werkzeug.utils import secure_filename
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
from pathlib import Path
from calendar import monthrange

app = Flask(__name__)
app.secret_key = 'dlocal-pending-debits-secret-key-2025'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create necessary folders
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_month_from_filename(filename):
    """Extract month name and year from filename"""
    month_mapping = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'may': 5, 'june': 6, 'july': 7, 'august': 8,
        'september': 9, 'october': 10, 'november': 11, 'december': 12
    }
    
    filename_lower = filename.lower()
    for month_name, month_num in month_mapping.items():
        if month_name in filename_lower:
            year_match = re.search(r'20\d{2}', filename)
            year = int(year_match.group()) if year_match else datetime.now().year
            return month_num, year
    
    return None, None


def get_following_month(month, year):
    """Get the following month number and year"""
    if month == 12:
        return 1, year + 1
    return month + 1, year


def parse_date(date_value):
    """Parse date from various formats"""
    if isinstance(date_value, datetime):
        return date_value
    if isinstance(date_value, str):
        for fmt in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(date_value, fmt)
            except:
                continue
    return None


def process_transaction_file(filepath):
    """Process the transaction file and calculate pending debits"""
    
    filename = Path(filepath).name
    file_month, file_year = extract_month_from_filename(filename)
    
    if not file_month:
        return None, "Could not extract month from filename"
    
    following_month, following_year = get_following_month(file_month, file_year)
    
    # Load workbook
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Find header row and column indices
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value:
                header_text = str(cell.value).strip()
                if 'Date processed' in header_text or 'Date Processed' in header_text:
                    headers['date_processed'] = cell.column
                elif 'ACH_DEBIT_AMOUNT' in header_text:
                    headers['debit_amount'] = cell.column
                elif 'ACH_RETURN_AMOUNT' in header_text:
                    headers['return_amount'] = cell.column
                elif header_text == 'Date':
                    headers['date'] = cell.column
        
        if len(headers) >= 3:
            header_row = cell.row
            break
    
    if 'date_processed' not in headers:
        wb.close()
        return None, "Could not find 'Date processed' column"
    
    # Find CN and DN columns
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value:
                header_text = str(cell.value).strip()
                if header_text == 'CN':
                    headers['cn'] = cell.column
                elif header_text == 'DN':
                    headers['dn'] = cell.column
    
    # Process transactions
    total_debit = 0
    total_return = 0
    filtered_transactions = []
    all_transactions = []
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        date_processed_cell = row[headers['date_processed'] - 1]
        date_processed = parse_date(date_processed_cell.value)
        
        if not date_processed_cell.value:
            continue
        
        debit_cell = row[headers['debit_amount'] - 1]
        debit_value = debit_cell.value
        debit_amount = 0
        if debit_value and debit_value != '-':
            try:
                debit_amount = float(str(debit_value).replace(',', ''))
            except:
                pass
        
        return_cell = row[headers['return_amount'] - 1]
        return_value = return_cell.value
        return_amount = 0
        if return_value and return_value != '-':
            try:
                return_amount = float(str(return_value).replace(',', ''))
            except:
                pass
        
        date_cell = row[headers.get('date', 0) - 1] if 'date' in headers else None
        cn_cell = row[headers.get('cn', 0) - 1] if 'cn' in headers else None
        dn_cell = row[headers.get('dn', 0) - 1] if 'dn' in headers else None
        
        is_following_month = date_processed and date_processed.month == following_month and date_processed.year == following_year
        
        all_transactions.append({
            'date': date_cell.value if date_cell else '',
            'date_processed': date_processed_cell.value,
            'debit': debit_value if debit_value and debit_value != '-' else 0,
            'return': return_value if return_value and return_value != '-' else 0,
            'cn': cn_cell.value if cn_cell else '',
            'dn': dn_cell.value if dn_cell else '',
            'is_following_month': is_following_month
        })
        
        if is_following_month:
            total_debit += debit_amount
            total_return += return_amount
            
            filtered_transactions.append({
                'date': date_cell.value if date_cell else '',
                'date_processed': date_processed_cell.value,
                'debit': debit_value if debit_value and debit_value != '-' else 0,
                'return': return_value if return_value and return_value != '-' else 0
            })
    
    net_amount = total_debit - total_return
    
    result = {
        'file_month': file_month,
        'file_year': file_year,
        'following_month': following_month,
        'following_year': following_year,
        'total_debit': total_debit,
        'total_return': total_return,
        'net_amount': net_amount,
        'transactions': filtered_transactions,
        'all_transactions': all_transactions,
        'filename': filename
    }
    
    wb.close()
    return result, None


def create_summary_and_je(result, output_filepath):
    """Create summary table and journal entry in a new Excel file"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "dLocal Pending Summary"
    
    # Styles
    header_font = Font(name='Calibri', size=11, bold=True, color='000000')
    normal_font = Font(name='Calibri', size=11, color='000000')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title
    ws['A1'] = 'dLocal Pending Debits Processing'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:F1')
    
    # File information
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    
    ws['A3'] = f"File Period: {month_names[result['file_month']]} {result['file_year']}"
    ws['A3'].font = Font(name='Calibri', size=11, bold=True)
    
    ws['A4'] = f"Date Processed Filter: {month_names[result['following_month']]} {result['following_year']}"
    ws['A4'].font = Font(name='Calibri', size=11, bold=True)
    
    # All Transactions Table
    row = 6
    ws[f'A{row}'] = 'ALL TRANSACTIONS'
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    headers_detail = ['Date', 'ACH_DEBIT_AMOUNT', 'ACH_RETURN_AMOUNT', 'Date processed', 'CN', 'DN', 'Net Amount (B-C)']
    for col_idx, header in enumerate(headers_detail, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align
    
    # Highlight color for following month rows
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Add ALL transaction rows
    for txn in result['all_transactions']:
        row += 1
        ws.cell(row=row, column=1).value = txn['date']
        
        debit_val = txn['debit'] if txn['debit'] != 0 else None
        return_val = txn['return'] if txn['return'] != 0 else None
        
        ws.cell(row=row, column=2).value = debit_val
        ws.cell(row=row, column=3).value = return_val
        ws.cell(row=row, column=4).value = txn['date_processed']
        ws.cell(row=row, column=5).value = txn.get('cn', '')
        ws.cell(row=row, column=6).value = txn.get('dn', '')
        
        # Add formula for Net Amount
        if txn.get('is_following_month', False):
            formula = f'=IFERROR(B{row}-C{row},"")'
            ws.cell(row=row, column=7).value = formula
        else:
            ws.cell(row=row, column=7).value = ''
        
        for col_idx in range(1, 8):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            
            if txn.get('is_following_month', False):
                cell.fill = highlight_fill
            
            if col_idx in [2, 3] and cell.value is not None:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            elif col_idx == 7:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
    
    # Summary Table
    row += 2
    ws[f'A{row}'] = 'SUMMARY'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].border = border_style
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = 'Description'
    ws[f'B{row}'] = 'Amount'
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
        ws[f'{col}{row}'].border = border_style
        ws[f'{col}{row}'].alignment = center_align
    
    row += 1
    ws[f'A{row}'] = 'Total ACH Debit Amount'
    ws[f'B{row}'] = result['total_debit']
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].border = border_style
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'B{row}'].alignment = right_align
    
    row += 1
    ws[f'A{row}'] = 'Total ACH Return Amount'
    ws[f'B{row}'] = result['total_return']
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].border = border_style
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'B{row}'].alignment = right_align
    
    row += 1
    ws[f'A{row}'] = 'Net Amount (Debit - Return)'
    ws[f'B{row}'] = result['net_amount']
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row}'].border = border_style
    ws[f'A{row}'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    ws[f'B{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'B{row}'].border = border_style
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'B{row}'].alignment = right_align
    ws[f'B{row}'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    
    # Journal Entry
    row += 3
    ws[f'A{row}'] = 'JOURNAL ENTRY'
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{row}:L{row}')
    
    row += 1
    je_headers = ['Debit', 'Credit', 'Date', 'Reversal Date', 'Memo', 'Account', 
                  'Department', 'Location', 'Name', 'Subsidiary', 'Journal Entry : Memo', 'Class']
    for col_idx, header in enumerate(je_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align
    
    # Generate JE
    net = result['net_amount']
    file_month_str = f"{result['file_month']:02d}"
    file_year_str = str(result['file_year'])
    memo = f"dLocal Pending Debits {file_month_str}.{file_year_str}"
    
    last_day = monthrange(result['file_year'], result['file_month'])[1]
    date_str = f"{result['file_month']}/{last_day}/{result['file_year']}"
    reversal_date_str = f"{result['following_month']}/1/{result['following_year']}"
    
    if net > 0:
        # Positive: Debit 22010, Credit 21017
        row += 1
        ws.cell(row=row, column=1).value = abs(net)
        ws.cell(row=row, column=2).value = None
        ws.cell(row=row, column=3).value = date_str
        ws.cell(row=row, column=4).value = reversal_date_str
        ws.cell(row=row, column=5).value = memo
        ws.cell(row=row, column=6).value = '22010 - Customer Funds Obligation : Customer Funds Liability'
        ws.cell(row=row, column=7).value = '0000 Corporate'
        ws.cell(row=row, column=8).value = '1 San Francisco'
        ws.cell(row=row, column=9).value = None
        ws.cell(row=row, column=10).value = 'Gusto Inc Global : Gusto Inc US'
        ws.cell(row=row, column=11).value = memo
        ws.cell(row=row, column=12).value = '601 Horizontal'
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            if col_idx in [1, 2]:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
        
        row += 1
        ws.cell(row=row, column=1).value = None
        ws.cell(row=row, column=2).value = abs(net)
        ws.cell(row=row, column=3).value = date_str
        ws.cell(row=row, column=4).value = reversal_date_str
        ws.cell(row=row, column=5).value = memo
        ws.cell(row=row, column=6).value = '21017 - Other Current Liabilities : Accrued Liabilities - Platform'
        ws.cell(row=row, column=7).value = '0000 Corporate'
        ws.cell(row=row, column=8).value = '1 San Francisco'
        ws.cell(row=row, column=9).value = None
        ws.cell(row=row, column=10).value = 'Gusto Inc Global : Gusto Inc US'
        ws.cell(row=row, column=11).value = memo
        ws.cell(row=row, column=12).value = '601 Horizontal'
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            if col_idx in [1, 2]:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
    
    elif net < 0:
        # Negative: Credit 22010, Debit 21017
        row += 1
        ws.cell(row=row, column=1).value = None
        ws.cell(row=row, column=2).value = abs(net)
        ws.cell(row=row, column=3).value = date_str
        ws.cell(row=row, column=4).value = reversal_date_str
        ws.cell(row=row, column=5).value = memo
        ws.cell(row=row, column=6).value = '22010 - Customer Funds Obligation : Customer Funds Liability'
        ws.cell(row=row, column=7).value = '0000 Corporate'
        ws.cell(row=row, column=8).value = '1 San Francisco'
        ws.cell(row=row, column=9).value = None
        ws.cell(row=row, column=10).value = 'Gusto Inc Global : Gusto Inc US'
        ws.cell(row=row, column=11).value = memo
        ws.cell(row=row, column=12).value = '601 Horizontal'
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            if col_idx in [1, 2]:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
        
        row += 1
        ws.cell(row=row, column=1).value = abs(net)
        ws.cell(row=row, column=2).value = None
        ws.cell(row=row, column=3).value = date_str
        ws.cell(row=row, column=4).value = reversal_date_str
        ws.cell(row=row, column=5).value = memo
        ws.cell(row=row, column=6).value = '21017 - Other Current Liabilities : Accrued Liabilities - Platform'
        ws.cell(row=row, column=7).value = '0000 Corporate'
        ws.cell(row=row, column=8).value = '1 San Francisco'
        ws.cell(row=row, column=9).value = None
        ws.cell(row=row, column=10).value = 'Gusto Inc Global : Gusto Inc US'
        ws.cell(row=row, column=11).value = memo
        ws.cell(row=row, column=12).value = '601 Horizontal'
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            if col_idx in [1, 2]:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
    
    else:
        row += 1
        ws[f'A{row}'] = 'No journal entry required (Net Amount = 0)'
        ws[f'A{row}'].font = Font(name='Calibri', size=11, italic=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 35
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 35
    ws.column_dimensions['L'].width = 15
    
    wb.save(output_filepath)
    wb.close()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('index'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('index'))
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        # Process the file
        result, error = process_transaction_file(filepath)
        
        if error:
            flash(f'Error processing file: {error}', 'error')
            os.remove(filepath)
            return redirect(url_for('index'))
        
        # Generate output file
        output_filename = f"{timestamp}_{filename.replace('.xlsx', '_Summary_JE.xlsx')}"
        output_filepath = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        
        create_summary_and_je(result, output_filepath)
        
        # Clean up uploaded file
        os.remove(filepath)
        
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']
        
        # Calculate last day of file month
        last_day = monthrange(result['file_year'], result['file_month'])[1]
        
        return render_template('result.html',
                             filename=filename,
                             output_filename=output_filename,
                             file_month=result['file_month'],
                             file_year=result['file_year'],
                             following_month=result['following_month'],
                             following_year=result['following_year'],
                             total_debit=result['total_debit'],
                             total_return=result['total_return'],
                             net_amount=result['net_amount'],
                             transaction_count=len(result['transactions']),
                             all_transactions=result['all_transactions'],
                             last_day=last_day)
    else:
        flash('Invalid file type. Please upload an Excel file (.xlsx or .xls)', 'error')
        return redirect(url_for('index'))


@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    return send_file(filepath, as_attachment=True)


@app.route('/download-je/<filename>')
def download_je_only(filename):
    """Generate and download JE template only (without transaction details)"""
    full_filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    
    # Load the full workbook
    wb_full = openpyxl.load_workbook(full_filepath)
    ws_full = wb_full.active
    
    # Create new workbook for JE only
    wb_je = openpyxl.Workbook()
    ws_je = wb_je.active
    ws_je.title = "Bank Fees Accrual JE Output"
    
    # Find where the JE section starts in the full workbook
    je_start_row = None
    for row in ws_full.iter_rows(min_row=1, values_only=False):
        if row[0].value and 'JOURNAL ENTRY' in str(row[0].value):
            je_start_row = row[0].row + 1  # Start from header row
            break
    
    if je_start_row:
        # Copy JE header and data rows
        current_je_row = 1
        for row_idx in range(je_start_row, ws_full.max_row + 1):
            row = ws_full[row_idx]
            # Stop if we hit an empty row after JE data
            if all(cell.value is None for cell in row[:12]):
                break
            
            # Copy cell values and formatting
            for col_idx in range(1, 13):  # Columns A-L
                source_cell = ws_full.cell(row=row_idx, column=col_idx)
                target_cell = ws_je.cell(row=current_je_row, column=col_idx)
                
                # Copy value
                target_cell.value = source_cell.value
                
                # Copy formatting
                if source_cell.font:
                    target_cell.font = source_cell.font.copy()
                if source_cell.border:
                    target_cell.border = source_cell.border.copy()
                if source_cell.fill:
                    target_cell.fill = source_cell.fill.copy()
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.alignment:
                    target_cell.alignment = source_cell.alignment.copy()
            
            current_je_row += 1
        
        # Copy column widths
        for col_idx in range(1, 13):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if col_letter in ws_full.column_dimensions:
                ws_je.column_dimensions[col_letter].width = ws_full.column_dimensions[col_letter].width
    
    wb_full.close()
    
    # Save JE-only workbook
    je_filename = filename.replace('_Summary_JE.xlsx', '_JE_Only.xlsx')
    je_filepath = os.path.join(app.config['OUTPUT_FOLDER'], je_filename)
    wb_je.save(je_filepath)
    wb_je.close()
    
    return send_file(je_filepath, as_attachment=True)


if __name__ == '__main__':
    print("\n" + "="*60)
    print("🚀 dLocal Pending Debits Processor")
    print("="*60)
    print("\n📍 Access the app at:")
    print("   👉 http://localhost:8080")
    print("   👉 http://127.0.0.1:8080")
    print("\n💡 Press CTRL+C to stop the server")
    print("="*60 + "\n")
    app.run(debug=True, host='0.0.0.0', port=8080)
