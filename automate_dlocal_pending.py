#!/usr/bin/env python3
"""
Automate dLocal Pending Debits Processing
Reads transaction file, filters by following month, and generates journal entries
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import re
import sys
from pathlib import Path


def extract_month_from_filename(filename):
    """Extract month name and year from filename like '09 Control Gusto Inc Septiembre 2025.xlsx'"""
    month_mapping = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'may': 5, 'june': 6, 'july': 7, 'august': 8,
        'september': 9, 'october': 10, 'november': 11, 'december': 12
    }
    
    filename_lower = filename.lower()
    for month_name, month_num in month_mapping.items():
        if month_name in filename_lower:
            # Extract year
            year_match = re.search(r'20\d{2}', filename)
            year = int(year_match.group()) if year_match else datetime.now().year
            return month_num, year
    
    return None, None


def get_following_month(month, year):
    """Get the following month number and year"""
    if month == 12:
        return 1, year + 1
    return month + 1, year


def parse_date(date_value):
    """Parse date from various formats"""
    if isinstance(date_value, datetime):
        return date_value
    if isinstance(date_value, str):
        # Try different date formats
        for fmt in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(date_value, fmt)
            except:
                continue
    return None


def process_transaction_file(filepath):
    """Process the transaction file and calculate pending debits"""
    
    # Extract month from filename
    filename = Path(filepath).name
    file_month, file_year = extract_month_from_filename(filename)
    
    if not file_month:
        print(f"Error: Could not extract month from filename: {filename}")
        return None
    
    following_month, following_year = get_following_month(file_month, file_year)
    
    print(f"Processing file for: {file_month}/{file_year}")
    print(f"Filtering for Date Processed in: {following_month}/{following_year}")
    
    # Load workbook
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Find header row and column indices
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value:
                header_text = str(cell.value).strip()
                if 'Date processed' in header_text or 'Date Processed' in header_text:
                    headers['date_processed'] = cell.column
                elif 'ACH_DEBIT_AMOUNT' in header_text:
                    headers['debit_amount'] = cell.column
                elif 'ACH_RETURN_AMOUNT' in header_text:
                    headers['return_amount'] = cell.column
                elif header_text == 'Date':
                    headers['date'] = cell.column
        
        if len(headers) >= 3:
            header_row = cell.row
            break
    
    if 'date_processed' not in headers:
        print("Error: Could not find 'Date processed' column")
        return None
    
    # Find CN and DN columns
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value:
                header_text = str(cell.value).strip()
                if header_text == 'CN':
                    headers['cn'] = cell.column
                elif header_text == 'DN':
                    headers['dn'] = cell.column
    
    # Process transactions
    total_debit = 0
    total_return = 0
    filtered_transactions = []
    all_transactions = []
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        date_processed_cell = row[headers['date_processed'] - 1]
        date_processed = parse_date(date_processed_cell.value)
        
        # Skip empty rows
        if not date_processed_cell.value:
            continue
        
        # Get debit amount
        debit_cell = row[headers['debit_amount'] - 1]
        debit_value = debit_cell.value
        debit_amount = 0
        if debit_value and debit_value != '-':
            try:
                debit_amount = float(str(debit_value).replace(',', ''))
            except:
                pass
        
        # Get return amount
        return_cell = row[headers['return_amount'] - 1]
        return_value = return_cell.value
        return_amount = 0
        if return_value and return_value != '-':
            try:
                return_amount = float(str(return_value).replace(',', ''))
            except:
                pass
        
        # Get other fields
        date_cell = row[headers.get('date', 0) - 1] if 'date' in headers else None
        cn_cell = row[headers.get('cn', 0) - 1] if 'cn' in headers else None
        dn_cell = row[headers.get('dn', 0) - 1] if 'dn' in headers else None
        
        # Check if this transaction is in the following month
        is_following_month = date_processed and date_processed.month == following_month and date_processed.year == following_year
        
        # Store ALL transactions
        all_transactions.append({
            'date': date_cell.value if date_cell else '',
            'date_processed': date_processed_cell.value,
            'debit': debit_value if debit_value and debit_value != '-' else 0,
            'return': return_value if return_value and return_value != '-' else 0,
            'cn': cn_cell.value if cn_cell else '',
            'dn': dn_cell.value if dn_cell else '',
            'is_following_month': is_following_month
        })
        
        # Check if this transaction should be included in the calculation
        if is_following_month:
            total_debit += debit_amount
            total_return += return_amount
            
            filtered_transactions.append({
                'date': date_cell.value if date_cell else '',
                'date_processed': date_processed_cell.value,
                'debit': debit_value if debit_value and debit_value != '-' else 0,
                'return': return_value if return_value and return_value != '-' else 0
            })
    
    net_amount = total_debit - total_return
    
    result = {
        'file_month': file_month,
        'file_year': file_year,
        'following_month': following_month,
        'following_year': following_year,
        'total_debit': total_debit,
        'total_return': total_return,
        'net_amount': net_amount,
        'transactions': filtered_transactions,
        'all_transactions': all_transactions,
        'filename': filename
    }
    
    wb.close()
    return result


def create_summary_and_je(result, output_filepath):
    """Create summary table and journal entry in a new Excel file"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "dLocal Pending Summary"
    
    # Styles
    header_font = Font(name='Calibri', size=11, bold=True, color='000000')
    normal_font = Font(name='Calibri', size=11, color='000000')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title
    ws['A1'] = 'dLocal Pending Debits Processing'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:F1')
    
    # File information
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    
    ws['A3'] = f"File Period: {month_names[result['file_month']]} {result['file_year']}"
    ws['A3'].font = Font(name='Calibri', size=11, bold=True)
    
    ws['A4'] = f"Date Processed Filter: {month_names[result['following_month']]} {result['following_year']}"
    ws['A4'].font = Font(name='Calibri', size=11, bold=True)
    
    # All Transactions Table
    row = 6
    ws[f'A{row}'] = 'ALL TRANSACTIONS'
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    headers_detail = ['Date', 'ACH_DEBIT_AMOUNT', 'ACH_RETURN_AMOUNT', 'Date processed', 'CN', 'DN', 'Net Amount (B-C)']
    for col_idx, header in enumerate(headers_detail, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align
    
    # Highlight color for following month rows
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
    
    # Add ALL transaction rows from result
    for txn in result['all_transactions']:
        row += 1
        ws.cell(row=row, column=1).value = txn['date']
        
        # For columns B and C, use actual values or None (not string '-')
        debit_val = txn['debit'] if txn['debit'] != 0 else None
        return_val = txn['return'] if txn['return'] != 0 else None
        
        ws.cell(row=row, column=2).value = debit_val
        ws.cell(row=row, column=3).value = return_val
        ws.cell(row=row, column=4).value = txn['date_processed']
        ws.cell(row=row, column=5).value = txn.get('cn', '')
        ws.cell(row=row, column=6).value = txn.get('dn', '')
        
        # Add formula for Net Amount (B - C) for rows with following month date processed
        if txn.get('is_following_month', False):
            # Use Excel formula: =IFERROR(B{row}-C{row},"")
            # This handles cases where B or C might be empty
            formula = f'=IFERROR(B{row}-C{row},"")'
            ws.cell(row=row, column=7).value = formula
        else:
            ws.cell(row=row, column=7).value = ''
        
        for col_idx in range(1, 8):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            
            # Highlight rows with following month date processed
            if txn.get('is_following_month', False):
                cell.fill = highlight_fill
            
            if col_idx in [2, 3] and cell.value is not None:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            elif col_idx == 7:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
    
    # Summary Table
    row += 2
    ws[f'A{row}'] = 'SUMMARY'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].border = border_style
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = 'Description'
    ws[f'B{row}'] = 'Amount'
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
        ws[f'{col}{row}'].border = border_style
        ws[f'{col}{row}'].alignment = center_align
    
    row += 1
    ws[f'A{row}'] = 'Total ACH Debit Amount'
    ws[f'B{row}'] = result['total_debit']
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].border = border_style
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'B{row}'].alignment = right_align
    
    row += 1
    ws[f'A{row}'] = 'Total ACH Return Amount'
    ws[f'B{row}'] = result['total_return']
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].border = border_style
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'B{row}'].alignment = right_align
    
    row += 1
    ws[f'A{row}'] = 'Net Amount (Debit - Return)'
    ws[f'B{row}'] = result['net_amount']
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row}'].border = border_style
    ws[f'A{row}'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    ws[f'B{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'B{row}'].border = border_style
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'B{row}'].alignment = right_align
    ws[f'B{row}'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    
    # Journal Entry
    row += 3
    ws[f'A{row}'] = 'JOURNAL ENTRY'
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{row}:L{row}')
    
    row += 1
    je_headers = ['Debit', 'Credit', 'Date', 'Reversal Date', 'Memo', 'Account', 
                  'Department', 'Location', 'Name', 'Subsidiary', 'Journal Entry : Memo', 'Class']
    for col_idx, header in enumerate(je_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align
    
    # Determine JE based on net amount
    net = result['net_amount']
    file_month_str = f"{result['file_month']:02d}"
    file_year_str = str(result['file_year'])
    memo = f"dLocal Pending Debits {file_month_str}.{file_year_str}"
    
    # Get the last day of the file month
    if result['file_month'] == 12:
        last_day_month = 12
        last_day_year = result['file_year']
    else:
        last_day_month = result['file_month']
        last_day_year = result['file_year']
    
    # Calculate last day
    from calendar import monthrange
    last_day = monthrange(last_day_year, last_day_month)[1]
    date_str = f"{last_day_month}/{last_day}/{last_day_year}"
    
    # Reversal date (first day of following month)
    reversal_date_str = f"{result['following_month']}/1/{result['following_year']}"
    
    if net > 0:
        # Positive: Debit Customer Funds Obligation, Credit Other Current Liabilities
        # Line 1: Debit Customer Funds Obligation (Debit column A)
        row += 1
        ws.cell(row=row, column=1).value = abs(net)  # Debit
        ws.cell(row=row, column=2).value = None  # Credit (empty)
        ws.cell(row=row, column=3).value = date_str
        ws.cell(row=row, column=4).value = reversal_date_str
        ws.cell(row=row, column=5).value = memo
        ws.cell(row=row, column=6).value = '22010 - Customer Funds Obligation : Customer Funds Liability'
        ws.cell(row=row, column=7).value = '0000 Corporate'
        ws.cell(row=row, column=8).value = '1 San Francisco'
        ws.cell(row=row, column=9).value = None  # Name is empty
        ws.cell(row=row, column=10).value = 'Gusto Inc Global : Gusto Inc US'
        ws.cell(row=row, column=11).value = memo
        ws.cell(row=row, column=12).value = '601 Horizontal'
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            if col_idx in [1, 2]:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
        
        # Line 2: Credit Other Current Liabilities (Credit column B)
        row += 1
        ws.cell(row=row, column=1).value = None  # Debit (empty)
        ws.cell(row=row, column=2).value = abs(net)  # Credit
        ws.cell(row=row, column=3).value = date_str
        ws.cell(row=row, column=4).value = reversal_date_str
        ws.cell(row=row, column=5).value = memo
        ws.cell(row=row, column=6).value = '21017 - Other Current Liabilities : Accrued Liabilities - Platform'
        ws.cell(row=row, column=7).value = '0000 Corporate'
        ws.cell(row=row, column=8).value = '1 San Francisco'
        ws.cell(row=row, column=9).value = None  # Name is empty
        ws.cell(row=row, column=10).value = 'Gusto Inc Global : Gusto Inc US'
        ws.cell(row=row, column=11).value = memo
        ws.cell(row=row, column=12).value = '601 Horizontal'
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            if col_idx in [1, 2]:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
    
    elif net < 0:
        # Negative: Credit Customer Funds Obligation, Debit Other Current Liabilities
        # Line 1: Credit Customer Funds Obligation (Credit column B)
        row += 1
        ws.cell(row=row, column=1).value = None  # Debit (empty)
        ws.cell(row=row, column=2).value = abs(net)  # Credit
        ws.cell(row=row, column=3).value = date_str
        ws.cell(row=row, column=4).value = reversal_date_str
        ws.cell(row=row, column=5).value = memo
        ws.cell(row=row, column=6).value = '22010 - Customer Funds Obligation : Customer Funds Liability'
        ws.cell(row=row, column=7).value = '0000 Corporate'
        ws.cell(row=row, column=8).value = '1 San Francisco'
        ws.cell(row=row, column=9).value = None  # Name is empty
        ws.cell(row=row, column=10).value = 'Gusto Inc Global : Gusto Inc US'
        ws.cell(row=row, column=11).value = memo
        ws.cell(row=row, column=12).value = '601 Horizontal'
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            if col_idx in [1, 2]:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
        
        # Line 2: Debit Other Current Liabilities (Debit column A)
        row += 1
        ws.cell(row=row, column=1).value = abs(net)  # Debit
        ws.cell(row=row, column=2).value = None  # Credit (empty)
        ws.cell(row=row, column=3).value = date_str
        ws.cell(row=row, column=4).value = reversal_date_str
        ws.cell(row=row, column=5).value = memo
        ws.cell(row=row, column=6).value = '21017 - Other Current Liabilities : Accrued Liabilities - Platform'
        ws.cell(row=row, column=7).value = '0000 Corporate'
        ws.cell(row=row, column=8).value = '1 San Francisco'
        ws.cell(row=row, column=9).value = None  # Name is empty
        ws.cell(row=row, column=10).value = 'Gusto Inc Global : Gusto Inc US'
        ws.cell(row=row, column=11).value = memo
        ws.cell(row=row, column=12).value = '601 Horizontal'
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = border_style
            if col_idx in [1, 2]:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
    
    else:
        # Zero amount
        row += 1
        ws[f'A{row}'] = 'No journal entry required (Net Amount = 0)'
        ws[f'A{row}'].font = Font(name='Calibri', size=11, italic=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 20  # ACH_DEBIT_AMOUNT
    ws.column_dimensions['C'].width = 20  # ACH_RETURN_AMOUNT
    ws.column_dimensions['D'].width = 18  # Date processed
    ws.column_dimensions['E'].width = 15  # CN
    ws.column_dimensions['F'].width = 15  # DN
    ws.column_dimensions['G'].width = 20  # Net Amount (B-C)
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 35
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 35
    ws.column_dimensions['L'].width = 15
    
    # Save workbook
    wb.save(output_filepath)
    wb.close()
    
    print(f"\nSummary and Journal Entry saved to: {output_filepath}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python automate_dlocal_pending.py <input_excel_file>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    if not Path(input_file).exists():
        print(f"Error: File not found: {input_file}")
        sys.exit(1)
    
    print("=" * 80)
    print("dLocal Pending Debits Automation")
    print("=" * 80)
    
    # Process the transaction file
    result = process_transaction_file(input_file)
    
    if not result:
        print("Error processing file")
        sys.exit(1)
    
    # Display results
    print(f"\nTotal ACH Debit Amount: ${result['total_debit']:,.2f}")
    print(f"Total ACH Return Amount: ${result['total_return']:,.2f}")
    print(f"Net Amount: ${result['net_amount']:,.2f}")
    print(f"Number of transactions filtered: {len(result['transactions'])}")
    
    # Generate output filename
    output_file = input_file.replace('.xlsx', '_Summary_JE.xlsx')
    
    # Create summary and JE
    create_summary_and_je(result, output_file)
    
    print("\n" + "=" * 80)
    print("Processing Complete!")
    print("=" * 80)


if __name__ == "__main__":
    main()
